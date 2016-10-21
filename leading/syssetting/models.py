from pymongo import ASCENDING, DESCENDING
from leading.config import leadingdb,leadingfiledb, DATABASE_DOMAIN, DATABASE_PORT,APPLICATION_DATA
from datetime import datetime
import subprocess
from flask import json
from openpyxl import Workbook,load_workbook
import os
import gridfs
from bson.objectid import ObjectId

class DatabaseBackup():
    def __init__(self):
        self.sourcedb_host = DATABASE_DOMAIN
        self.sourcedb_port = str(DATABASE_PORT)
        self.backupdb_host = "localhost"
        self.backupdb_port = "27017"
        self.backup_databse = ['leadingdb', 'leadingfiledb']
        self.createDate = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.temp_folder = '/tmp/'
        self.download_folder = '/tmp/'

    def backup(self, **kwargs):
        subprocess.check_call(['/bin/mkdir', '-p', self.temp_folder + 'backup'])
        # subprocess.check_call(['sudo','chown', 'mvrt', 'backup'])
        filenames = []
        for db in self.backup_databse:
            subprocess.check_call(
                ['/usr/local/bin/mongodump', "--host", self.sourcedb_host, "--port", self.sourcedb_port, "-d", db,
                 "--archive=" + self.temp_folder + "backup/" + db + "_backup"])
            subprocess.check_call(
                ['/usr/local/bin/mongofiles', "--host", self.backupdb_host, "--port", self.backupdb_port, "-d",
                 "dbbackup",
                 "--local",
                 self.temp_folder + "backup/" + db + "_backup", "--replace", "put", db + "_backup_" + self.createDate])
            filenames.append(db + "_backup_" + self.createDate)
        leadingdb.backup.update_one({"backupDate": self.createDate}, {
            "$set": {"backupName": "LEADING_BACKUP_" + self.createDate, "database": self.backup_databse,
                     "filenames": filenames, "backupHost": self.backupdb_host, "backupPort": self.backupdb_port,
                     "username": kwargs['username']}}, upsert=True)
        return self.get_list()

    def get_list(self):
        backups = leadingdb.backup.find({}, {"_id": 0}).sort("backupDate", DESCENDING)
        result = []
        for backup in backups:
            result.append(backup)
        return result

    def delete(self, data):
        data = json.loads(data)
        if len(data.keys()) > 0:
            if len(data['filenames']) > 0:
                for file in data['filenames']:
                    subprocess.check_call(
                        ['/usr/local/bin/mongofiles', "--host", data['backupHost'], "--port", data['backupPort'], "-d",
                         'dbbackup',
                         "delete", file])
                leadingdb.backup.delete_one({"backupDate": data['backupDate']})
        return self.get_list()

    def restore(self, data):
        data = json.loads(data)
        if len(data.keys()) > 0:
            if len(data['filenames']) > 0:
                subprocess.check_call(['/bin/mkdir', '-p', self.temp_folder + 'backup'])
                for file in data['filenames']:
                    subprocess.check_call(
                        ['/usr/local/bin/mongofiles', "--host", data['backupHost'], "--port", data['backupPort'],
                         "--local",
                         self.temp_folder + "backup/" + file, "-d", 'dbbackup', "get", file])
                    ss = subprocess.check_call(
                        ['/usr/local/bin/mongorestore', "--host", self.sourcedb_host, "--port", self.sourcedb_port,
                         "--drop",
                         "--archive=" + self.temp_folder + "backup/" + file])
        return self.get_list()

    def download(self, data):
        data = json.loads(data)
        if len(data.keys()) > 0:
            if len(data['filenames']) > 0:
                # subprocess.check_call(['/bin/mkdir', '-p', self.temp_folder + 'backup'])
                for file in data['filenames']:
                    subprocess.check_call(
                        ['/usr/local/bin/mongofiles', "--host", data['backupHost'], "--port", data['backupPort'],
                         "--local",
                         self.download_folder + file, "-d", 'dbbackup', "get", file])
        return self.get_list()


class SystemSetting():
    def __init__(self, settingId=None, userName=None):
        self.db = leadingdb
        self._id = settingId
        self.userName = userName

    def get_all(self):
        results = []
        result = self.db.settings.find().sort('group', ASCENDING)

        for res in result:
            res['_id'] = str(res['_id'])
            results.append(res)
        return results

    def update_id(self, **kwargs):
        exist_data = self.db.settings.find_one({'group': kwargs['group']})
        if exist_data != None:
            self._id = exist_data['_id']

    def set(self, **kwargs):
        self.update_id(group=kwargs['group'])
        self.db.settings.update_one({"_id": self._id}, {"$set": kwargs}, upsert=True)

    def get_system_current_period(self):
        systemInfo = self.db.systeminfo.find_one({"group": "systemInfo"}, {"_id": 0})
        return systemInfo['content'][0]['value']

    def upgrade_system_current_period(self):
        systemInfo = self.db.systeminfo.find_one({"group": "systemInfo"}, {"_id": 0})
        currentPeriod = systemInfo['content'][0]['value']
        self.db.systeminfo.update_one({"group": "systemInfo"}, {"$set": {"content.0.value": currentPeriod + 1}})
        return currentPeriod + 1

class DataInitialization():
    def __init__(self):
        self.db = leadingdb

    def set_data_config(self,dataConf):
        self.db.data_conf_def.update_one({'item':dataConf['item']},{"$set":dataConf},upsert=True)
        return self.get_data_config()

    def get_data_config(self):
        result = self.db.data_conf_def.find({},{"_id":0})
        return list(result)

    def save_file_tmp(self, file, filename, content_type):
        if os.path.isfile(os.path.join(APPLICATION_DATA,filename)):
            os.remove(os.path.join(APPLICATION_DATA,filename))
        f = open(os.path.join(APPLICATION_DATA,filename), 'w')
        f.write(file)
        f.close()

        # gfs = gridfs.GridFS(leadingfiledb)
        # objectID = gfs.put(file,content_type=content_type,filename=filename)
        # gridfsOut = gfs.get(objectID)  # .length,gfs.get(objectID).upload_date
        # output={"objectID": str(objectID),
        #                "content_type": gridfsOut.content_type[0],
        #                "filename": gridfsOut.filename,
        #                "length": gridfsOut.length,
        #                "upload_date": gridfsOut.upload_date.strftime('%Y-%m-%d %H:%M:%S')
        #                }
        return filename


    def init_db_from_excel(self,dataConf):
        wb = load_workbook(os.path.join(APPLICATION_DATA,dataConf['filename']))
        sheetsList = wb.get_sheet_names()
        if dataConf['sheetname'] not in sheetsList:
            return {'error': 'Can not find the Sheet Name in the file.'}
        else:
            ws = wb[dataConf['sheetname']]
            title = []
            for index, row in enumerate(ws.iter_rows()):
                if index == 0:
                    for cel in row:
                        title.append(cel.value)
                else:
                    rowDic = {}
                    for i, c in enumerate(row):
                        rowDic[title[i]] = c.value
                    self.db.get_collection(dataConf['sheetname']).update_one({dataConf['key']: rowDic[dataConf['key']]},
                                                                             {"$set": rowDic}, upsert=True)

            return {'status': "import data from " + dataConf['filename'] + ' successfully.'}

    def export_db_to_excel(self, dataConf):

        wb = Workbook()
        ws = wb.active
        ws.title = dataConf['sheetname']
        # ws = wb.create_sheet(dataConf['sheetname'])
        sortList = []
        for k in dataConf['key'].split(','):
            sortList.append((k, ASCENDING))
        source_data = self.db.get_collection(dataConf['sheetname']).find({}, {'_id': 0}).sort(sortList)
        if 'orderlist' in dataConf.keys():
            for i1, key1 in enumerate(dataConf['orderlist'].split(',')):
                ws.cell(row=1, column=i1 + 1, value=key1)

            for index, sd in enumerate(source_data):
                for i2, key2 in enumerate(dataConf['orderlist'].split(',')):
                    ws.cell(row=2 + index, column=i2 + 1, value=sd[key2])

        else:
            for index, sd in enumerate(source_data):
                if index == 0:
                    for i1, key1 in enumerate(sd.keys()):
                        ws.cell(row=1, column=i1 + 1, value=key1)
                for i2, key2 in enumerate(sd.keys()):
                    ws.cell(row=2 + index, column=i2 + 1, value=sd[key2])

        wb.save(os.path.join(APPLICATION_DATA, dataConf['filename']))
        return dataConf
