__author__ = 'isaac'
import os
import pymongo
from leading.config import leadingdb,APPLICATION_DATA,TESTING
from flask import json
from datetime import datetime
from bson.son import SON
from openpyxl import Workbook,load_workbook

class DatabaseInit():
    def __init__(self):
        self.db = leadingdb

    def import_from_json(self,dataFileList):
        for filename in dataFileList:
            if self.db.get_collection(filename).find().count() == 0:
                with open(os.path.join(APPLICATION_DATA, filename+'.json')) as data:
                    itemList = json.load(data)
                    data.close()
                    self.db.get_collection(filename).drop()
                    for item in itemList:
                        self.db.get_collection(filename).insert_one(item)

                print "Input Data from : " + filename + " Completed."


dataFileList = ['systeminfo', 'user', 'workflow', 'task_list', 'teams', 'periods_def', 'employees_def', 'workforce_def',
                'resources_def',
                'budget_def','actions_def','negotiation_def','niches_def','projects_def','corporate_acquisitions_def']
DatabaseInit().import_from_json(dataFileList)


# def initDB():
#     if sdb.sys_tasks_list.find().count() >= 0:
#         with open(os.path.join(APPLICATION_DATA, 'sys_task_list.json')) as data:
#             tasks = json.load(data)
#             data.close()
#             sdb.sys_tasks_list.drop()
#             for task in tasks:
#                 sdb.sys_tasks_list.insert_one(task)

    # if sdb.sys_periods_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'sys_periods_def.json')) as data:
    #         periods = json.load(data)
    #         data.close()
    #         sdb.sys_periods_def.drop()
    #         for period in periods:
    #             sdb.sys_periods_def.insert_one(period)
    #
    # if sdb.employees_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'employees_def.json')) as data:
    #         employees = json.load(data)
    #         data.close()
    #         sdb.employees_def.drop()
    #         for employee in employees:
    #             sdb.employees_def.insert_one(employee)
    #
    # if sdb.account_desc.find().count()== 0:
    #     with open(os.path.join(APPLICATION_DATA, 'account_desc.json')) as data:
    #         account_desc = json.load(data)
    #         data.close()
    #         sdb.account_desc.drop()
    #         for account in account_desc:
    #             sdb.account_desc.insert_one(account)
    #
    #
    # if sdb.budget_def.find().count()== 0:
    #     with open(os.path.join(APPLICATION_DATA, 'budget_def.json')) as data:
    #         budget_def = json.load(data)
    #         data.close()
    #         sdb.budget_def.drop()
    #         for budget in budget_def:
    #             sdb.budget_def.insert_one(budget)
    #
    # if sdb.account_ini.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'account_ini.json')) as data:
    #         account = json.load(data)
    #         data.close()
    #         sdb.account_ini.drop()
    #         for acc in account:
    #             sdb.account_ini.insert_one(acc)
    #
    # if sdb.teams_ini.find().count() >= 0:
    #     with open(os.path.join(APPLICATION_DATA, 'teams_ini.json')) as data:
    #         teams = json.load(data)
    #         data.close()
    #         sdb.teams_ini.drop()
    #         for team in teams:
    #             sdb.teams_ini.insert_one(team)
    #
    # if sdb.projects_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'projects_def.json')) as data:
    #         projects = json.load(data)
    #         data.close()
    #         sdb.projects_def.drop()
    #         for project in projects:
    #             sdb.projects_def.insert_one(project)
    #
    # if sdb.actions_def.find().count()== 0:
    #     with open(os.path.join(APPLICATION_DATA, 'actions_def.json')) as data:
    #         actions = json.load(data)
    #         data.close()
    #         sdb.actions_def.drop()
    #         for action in actions:
    #             sdb.actions_def.insert_one(action)
    #
    # if sdb.stories_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'stories_def.json')) as data:
    #         sidecolumnstories = json.load(data)
    #         data.close()
    #         sdb.stories_def.drop()
    #         for sidecolumnstory in sidecolumnstories:
    #             sdb.stories_def.insert_one(sidecolumnstory)
    #
    # if sdb.negotiation_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'negotiation_def.json')) as data:
    #         negotiationhr = json.load(data)
    #         data.close()
    #         sdb.negotiation_def.drop()
    #         for nhr in negotiationhr:
    #             sdb.negotiation_def.insert_one(nhr)
    #
    # if sdb.niches_def.find().count() == 0:
    #     with open(os.path.join(APPLICATION_DATA, 'niches_def.json')) as data:
    #         tmp_niches = json.load(data)
    #         data.close()
    #         sdb.niches_def.drop()
    #         for niche in tmp_niches:
    #             sdb.niches_def.insert_one(niche)
    #
    # if sdb.resources_def.find().count()== 0:
    #     with open(os.path.join(APPLICATION_DATA, 'resources_def.json')) as data:
    #         resources = json.load(data)
    #         data.close()
    #         sdb.resources_def.drop()
    #         for resource in resources:
    #             sdb.resources_def.insert_one(resource)
    #
    # if sdb.corporate_acquisitions_def.find().count()>=0:
    #     with open(os.path.join(APPLICATION_DATA, 'corporate_acquisitions_def.json')) as data:
    #         corporates = json.load(data)
    #         data.close()
    #         sdb.corporate_acquisitions_def.drop()
    #         for corporate in corporates:
    #             sdb.corporate_acquisitions_def.insert_one(corporate)




# def export_db_to_excel(list):
#
#     wb = Workbook()
#
#     ws0 = wb.active
#     ws0.title='task_list'
#     source_data = sdb.sys_tasks_list.find({'period':{"$gt":0}}, {"_id": 0,"preProcess":0}).sort([('companyName',pymongo.ASCENDING)
#                                                                                                     ,('period',pymongo.ASCENDING),('processNo',pymongo.ASCENDING)])
#     keys=['companyName','period','processNo','taskID','taskName','comment']
#     for index1,sd in enumerate(source_data):
#         if index1 == 0:
#             for index2,key in enumerate(keys):
#                 ws0.cell(column=index2+1, row=1, value=key)
#         for index3,key in enumerate(keys):
#             ws0.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'account_desc' in list:
#         ws1 = wb.create_sheet('account_desc')
#         keys1 =['accountDescType','accountDescLevel','accountDescID','accountDescName','accountDescription','summaryFLag',
#                 'subAccountName','calculateFlag']
#         source_data = sdb.account_desc.find({}, {"_id": 0}).sort('accountDescID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(keys1):
#                     ws1.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(keys1):
#                 ws1.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'employees' in list:
#         ws2 =wb.create_sheet('employees')
#         keys2 =['companyName','category','title','employeeID','employeeName','startAtPeriod','minimumSalary',
#                 'competenceIndexEffect','legitimacy','status','willingToMove']
#         source_data = sdb.employees_def.find({}, {"_id": 0,'ratingsInStars':0}).sort('employeeID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(keys2):
#                     ws2.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(keys2):
#                 ws2.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'actions' in list:
#         ws3 =wb.create_sheet('actions')
#         source_data = sdb.actions_def.find({}, {"_id": 0}).sort([('companyName',pymongo.ASCENDING),('category',pymongo.ASCENDING),('actionID',pymongo.ASCENDING)])
#         keys3=['companyName','category','actionID','actionName','periodStart','immediateIncrementalCost','COSChange',
#                'associatedCost','competenceIndex','stressIndex','adaptabilityIndex','legitimacy']
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(keys3):
#                     ws3.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(keys3):
#                 ws3.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'niches' in list:
#         ws4 =wb.create_sheet('niches')
#         source_data = sdb.niches_def.find({}, {"_id": 0}).sort('niche',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(sd.keys()):
#                     ws4.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(sd.keys()):
#                 ws4.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'negotiation' in list:
#         ws5 =wb.create_sheet('negotiation')
#         source_data = sdb.negotiation_def.find({}, {"_id": 0}).sort('employeeID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key0 in enumerate(sd.keys()):
#                     #if key0 == 'influenceBVs':
#                     #for index21,subkey21 in enumerate(sd[key0].keys()):
#                     #ws5.cell(column=len(sd.keys())+index21+1, row=1, value=subkey21)
#                     #elif key0 == 'technicalExperts':
#                     #for index22,subkey22 in enumerate(sd[key0].keys()):
#                     # ws5.cell(column=len(sd.keys())+index22+7, row=1, value=subkey22)
#                     ws5.cell(column=index2+1, row=1, value=key0)
#             for index3,key in enumerate(sd.keys()):
#                 #if key == 'influenceBVs':
#                 #for index4,subkey in enumerate(sd[key].keys()):
#                 #ws5.cell(column=len(sd.keys())+index4+1, row=index1+2, value=sd[key][subkey])
#                 #elif key == 'technicalExperts':
#                 #for index5,subkey2 in enumerate(sd[key].keys()):
#                 #ws5.cell(column=len(sd.keys())+index5+7, row=index1+2, value=sd[key][subkey2])
#                 ws5.cell(column=index3+1, row=index1+2, value=str(sd[key]))
#
#     if 'projects' in list:
#         ws6 =wb.create_sheet('projects')
#         source_data = sdb.projects_def.find({}, {"_id": 0}).sort([('Item#',pymongo.ASCENDING),('projectName',pymongo.ASCENDING)])
#         keys6 =['Item#','projectName','startAtPeriod','finalAtPeriod','lowerOfCost','strategicLogic','costHitPDbudget',
#                 'competence','stress','market','status']
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key0 in enumerate(keys6):
#                     ws6.cell(column=index2+1, row=1, value=key0)
#                     #if key0=='costHitPDbudget':
#                     #for index5,subkey in enumerate(sd[key0].keys()):
#                     #ws6.cell(column=len(sd.keys())+index5+1, row=1, value=subkey)
#             for index3,key in enumerate(keys6):
#                 #if key=='costHitPDbudget':
#                 #for index4,subkey in enumerate(sd[key].keys()):
#                 #ws6.cell(column=len(sd.keys())+index4+1, row=index1+2, value=sd[key][subkey])
#                 #else:
#                 ws6.cell(column=index3+1, row=index1+2, value=str(sd[key]))
#
#     if 'resources' in list:
#         ws7 =wb.create_sheet('resources')
#         source_data = sdb.resources_def.find({}, {"_id": 0}).sort('employeeID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(sd.keys()):
#                     ws7.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(sd.keys()):
#                 ws7.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'stories' in list:
#         ws8 =wb.create_sheet('stories')
#         source_data = sdb.stories_def.find({}, {"_id": 0}).sort('projectID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(sd.keys()):
#                     ws8.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(sd.keys()):
#                 ws8.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     if 'corporates' in list:
#         ws9 =wb.create_sheet('corporates')
#         source_data = sdb.corporate_acquisitions_def.find({}, {"_id": 0}).sort('companyID',pymongo.ASCENDING)
#         for index1,sd in enumerate(source_data):
#             if index1 == 0:
#                 for index2,key in enumerate(sd.keys()):
#                     ws9.cell(column=index2+1, row=1, value=key)
#             for index3,key in enumerate(sd.keys()):
#                 ws9.cell(column=index3+1, row=index1+2, value=sd[key])
#
#     wb.save(os.path.join(APPLICATION_DATA, 'init_data.xlsx'))
#     print 'export'
#
#
# def init_db_from_excel():
#     wb = load_workbook(os.path.join(APPLICATION_DATA, 'init_data.xlsx'))
#     print wb.get_sheet_names()
#     sheetsList = wb.get_sheet_names()
#     if 'account_desc' in sheetsList:
#         ws1 = wb['account_desc']
#         title=[]
#         for c in range(1,len(ws1.columns)+1):
#             title.append(ws1.cell(row=1,column=c).value)
#         for r in range(2,len(ws1.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws1.columns)+1):
#                 row_value[title[c-1]] = ws1.cell(row=r,column=c).value
#             sdb.account_desc.update_one({"accountDescID":row_value['accountDescID']},{"$set":row_value},upsert=True)
#
#     if 'employees' in sheetsList:
#         ws2 = wb['employees']
#         title=[]
#         for c in range(1,len(ws2.columns)+1):
#             title.append(ws2.cell(row=1,column=c).value)
#         for r in range(2,len(ws2.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws2.columns)+1):
#                 row_value[title[c-1]] = ws2.cell(row=r,column=c).value
#             sdb.employees_def.update_one({"employeeID":row_value['employeeID']},{"$set":row_value},upsert=True)
#
#     if 'actions' in sheetsList:
#         ws3 = wb['actions']
#         title=[]
#         for c in range(1,len(ws3.columns)+1):
#             title.append(ws3.cell(row=1,column=c).value)
#         for r in range(2,len(ws3.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws3.columns)+1):
#                 row_value[title[c-1]] = ws3.cell(row=r,column=c).value
#             sdb.actions_def.update_one({"actionID":row_value['actionID']},{"$set":row_value},upsert=True)
#
#     if 'niches' in sheetsList:
#         ws4 = wb['niches']
#         title=[]
#         for c in range(1,len(ws4.columns)+1):
#             title.append(ws4.cell(row=1,column=c).value)
#         for r in range(2,len(ws4.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws4.columns)+1):
#                 row_value[title[c-1]] = ws4.cell(row=r,column=c).value
#             sdb.niches_def.update_one({"niche":row_value['niche'],"period":row_value['period']},{"$set":row_value},upsert=True)
#
#     if 'negotiation' in sheetsList:
#         ws5 = wb['negotiation']
#         title=[]
#         for c in range(1,len(ws5.columns)+1):
#             title.append(ws5.cell(row=1,column=c).value)
#         for r in range(2,len(ws5.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws5.columns)+1):
#                 row_value[title[c-1]] = ws5.cell(row=r,column=c).value
#             sdb.negotiation_def.update_one({"employeeID":row_value['employeeID']},{"$set":row_value},upsert=True)
#
#     if 'projects' in sheetsList:
#         ws6 = wb['projects']
#         title=[]
#         for c in range(1,len(ws6.columns)+1):
#             title.append(ws6.cell(row=1,column=c).value)
#         for r in range(2,len(ws6.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws6.columns)+1):
#                 row_value[title[c-1]] = ws6.cell(row=r,column=c).value
#             sdb.projects_def.update_one({"projectName":row_value['projectName']},{"$set":row_value},upsert=True)
#
#     if 'resources' in sheetsList:
#         ws7 = wb['resources']
#         title=[]
#         for c in range(1,len(ws7.columns)+1):
#             title.append(ws7.cell(row=1,column=c).value)
#         for r in range(2,len(ws7.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws7.columns)+1):
#                 row_value[title[c-1]] = ws7.cell(row=r,column=c).value
#             sdb.resources_def.update_one({"resourceName":row_value['resourceName']},{"$set":row_value},upsert=True)
#
#     if 'stories' in sheetsList:
#         ws8 = wb['stories']
#         title=[]
#         for c in range(1,len(ws8.columns)+1):
#             title.append(ws8.cell(row=1,column=c).value)
#         for r in range(2,len(ws8.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws8.columns)+1):
#                 row_value[title[c-1]] = ws8.cell(row=r,column=c).value
#             sdb.stories_def.update_one({"storyID":row_value['storyID']},{"$set":row_value},upsert=True)
#
#     if 'corporates' in sheetsList:
#         ws9 = wb['corporates']
#         title=[]
#         for c in range(1,len(ws9.columns)+1):
#             title.append(ws9.cell(row=1,column=c).value)
#         for r in range(2,len(ws9.rows)+1):
#             row_value = {}
#             for c in range(1,len(ws9.columns)+1):
#                 row_value[title[c-1]] = ws9.cell(row=r,column=c).value
#             sdb.corporate_acquisitions_def.update_one({"companyID":row_value['companyID']},{"$set":row_value},upsert=True)
#
#
#     print 'init_excel'
#     #export_db_to_excel()
#     #init_db_from_excel()